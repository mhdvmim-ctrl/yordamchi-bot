import asyncio
import io
import json
import logging
import math
import os
from datetime import datetime, timedelta, timezone
from typing import Optional

from dotenv import load_dotenv
from openpyxl import Workbook
from sqlalchemy import (
    BigInteger,
    Boolean,
    DateTime,
    ForeignKey,
    Integer,
    String,
    Text,
    func,
    select,
    update,
)
from sqlalchemy.ext.asyncio import (
    AsyncAttrs,
    AsyncSession,
    async_sessionmaker,
    create_async_engine,
)
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, relationship

from aiogram import Bot, Dispatcher, F, Router
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ContentType, ParseMode
from aiogram.exceptions import TelegramBadRequest, TelegramForbiddenError
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    FSInputFile,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
)
from aiogram.utils.keyboard import ReplyKeyboardBuilder


# ============================================================
# CONFIGURATION
# ============================================================

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
ADMIN_ID_RAW = os.getenv("ADMIN_ID", "").strip()
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "mhdvmim").strip().lower()

if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN .env faylida ko'rsatilmagan.")

if not ADMIN_ID_RAW.isdigit():
    raise RuntimeError("ADMIN_ID .env faylida raqam bo'lishi kerak.")

ADMIN_ID = int(ADMIN_ID_RAW)

DATABASE_URL = "sqlite+aiosqlite:///bot_database.db"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)

logger = logging.getLogger("mhdv_bot")


# ============================================================
# DATABASE
# ============================================================

class Base(AsyncAttrs, DeclarativeBase):
    pass


class User(Base):
    __tablename__ = "users"

    telegram_id: Mapped[int] = mapped_column(BigInteger, primary_key=True)
    full_name: Mapped[str] = mapped_column(String(255))
    username: Mapped[Optional[str]] = mapped_column(String(255), nullable=True)
    is_blocked: Mapped[bool] = mapped_column(Boolean, default=False)
    discount_percent: Mapped[int] = mapped_column(Integer, default=0)
    created_at: Mapped[datetime] = mapped_column(
        DateTime,
        default=datetime.utcnow,
    )

    orders: Mapped[list["Order"]] = relationship(
        back_populates="user",
        cascade="all, delete-orphan",
    )

    feedbacks: Mapped[list["Feedback"]] = relationship(
        back_populates="user",
        cascade="all, delete-orphan",
    )


class Order(Base):
    __tablename__ = "orders"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    user_id: Mapped[int] = mapped_column(
        BigInteger,
        ForeignKey("users.telegram_id"),
        nullable=False,
    )
    order_type: Mapped[str] = mapped_column(String(20))
    details: Mapped[str] = mapped_column(Text)
    status: Mapped[str] = mapped_column(
        String(20),
        default="pending",
    )
    rejection_reason: Mapped[Optional[str]] = mapped_column(
        Text,
        nullable=True,
    )
    created_at: Mapped[datetime] = mapped_column(
        DateTime,
        default=datetime.utcnow,
    )

    user: Mapped["User"] = relationship(back_populates="orders")
    payments: Mapped[list["Payment"]] = relationship(
        back_populates="order",
        cascade="all, delete-orphan",
    )


class Payment(Base):
    __tablename__ = "payments"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    order_id: Mapped[int] = mapped_column(
        Integer,
        ForeignKey("orders.id"),
        nullable=False,
    )
    file_id: Mapped[str] = mapped_column(String(500))
    status: Mapped[str] = mapped_column(
        String(20),
        default="pending",
    )
    created_at: Mapped[datetime] = mapped_column(
        DateTime,
        default=datetime.utcnow,
    )

    order: Mapped["Order"] = relationship(back_populates="payments")


class Feedback(Base):
    __tablename__ = "feedbacks"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    user_id: Mapped[int] = mapped_column(
        BigInteger,
        ForeignKey("users.telegram_id"),
        nullable=False,
    )
    type: Mapped[str] = mapped_column(String(20))
    text: Mapped[str] = mapped_column(Text)
    created_at: Mapped[datetime] = mapped_column(
        DateTime,
        default=datetime.utcnow,
    )

    user: Mapped["User"] = relationship(back_populates="feedbacks")


class PromoCode(Base):
    __tablename__ = "promo_codes"

    code: Mapped[str] = mapped_column(String(100), primary_key=True)
    discount_percent: Mapped[int] = mapped_column(Integer)
    max_uses: Mapped[int] = mapped_column(Integer)
    used_count: Mapped[int] = mapped_column(Integer, default=0)
    expires_at: Mapped[datetime] = mapped_column(DateTime)


class DynamicContent(Base):
    __tablename__ = "dynamic_contents"

    key: Mapped[str] = mapped_column(String(100), primary_key=True)
    value: Mapped[str] = mapped_column(Text)


engine = create_async_engine(
    DATABASE_URL,
    echo=False,
    future=True,
)

SessionLocal = async_sessionmaker(
    engine,
    class_=AsyncSession,
    expire_on_commit=False,
)


# ============================================================
# DEFAULT CONTENT
# ============================================================

DEFAULT_CONTENT = {
    "about_bot": (
        "🤖 MHDV Bot\n\n"
        "MHDV xizmatlari uchun buyurtma berish, ma'lumot olish, "
        "portfolio ko'rish va administrator bilan bog'lanish imkonini beradi."
    ),
    "about_mhdv": (
        "🏢 MHDV\n\n"
        "MHDV — zamonaviy web saytlar, logo va raqamli xizmatlar "
        "yaratishga yo'naltirilgan loyiha."
    ),
    "card_info": (
        "💳 To'lov rekvizitlari\n\n"
        "Karta: 0000 0000 0000 0000\n"
        "Qabul qiluvchi: MHDV\n\n"
        "To'lovdan so'ng chekni ushbu botga yuboring."
    ),
    "socials": (
        "🌐 Ijtimoiy tarmoqlar\n\n"
        "Telegram: @mhdv_programer\n"
        "Web: MHDV\n"
        "Portfolio: MHDV"
    ),
    "ws_info": (
        "🌐 Veb sayt xizmatlari\n\n"
        "Landing page, korporativ sayt, portfolio, "
        "internet do'kon va maxsus web tizimlar ishlab chiqiladi."
    ),
    "logo_info": (
        "🎨 Logo xizmatlari\n\n"
        "Brend uchun zamonaviy, minimalistik va professional "
        "logo dizaynlar tayyorlanadi."
    ),
}


async def init_db():
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    async with SessionLocal() as session:
        for key, value in DEFAULT_CONTENT.items():
            result = await session.execute(
                select(DynamicContent).where(DynamicContent.key == key)
            )
            existing = result.scalar_one_or_none()

            if existing is None:
                session.add(
                    DynamicContent(
                        key=key,
                        value=value,
                    )
                )

        await session.commit()


# ============================================================
# HELPERS
# ============================================================

async def get_user(telegram_id: int) -> Optional[User]:
    async with SessionLocal() as session:
        result = await session.execute(
            select(User).where(User.telegram_id == telegram_id)
        )
        return result.scalar_one_or_none()


async def ensure_user(message: Message) -> Optional[User]:
    if not message.from_user:
        return None

    telegram_id = message.from_user.id

    async with SessionLocal() as session:
        result = await session.execute(
            select(User).where(User.telegram_id == telegram_id)
        )
        user = result.scalar_one_or_none()

        full_name = (
            message.from_user.full_name
            or message.from_user.first_name
            or "Foydalanuvchi"
        )

        username = message.from_user.username

        if user is None:
            user = User(
                telegram_id=telegram_id,
                full_name=full_name,
                username=username,
            )
            session.add(user)
        else:
            user.full_name = full_name
            user.username = username

        await session.commit()

        return user


async def get_dynamic_content(key: str) -> str:
    async with SessionLocal() as session:
        result = await session.execute(
            select(DynamicContent).where(DynamicContent.key == key)
        )
        item = result.scalar_one_or_none()

        if item:
            return item.value

        return DEFAULT_CONTENT.get(key, "")


async def set_dynamic_content(key: str, value: str):
    async with SessionLocal() as session:
        result = await session.execute(
            select(DynamicContent).where(DynamicContent.key == key)
        )

        item = result.scalar_one_or_none()

        if item is None:
            session.add(
                DynamicContent(
                    key=key,
                    value=value,
                )
            )
        else:
            item.value = value

        await session.commit()


async def is_blocked_user(telegram_id: int) -> bool:
    user = await get_user(telegram_id)

    return bool(user and user.is_blocked)


async def safe_send(
    bot: Bot,
    user_id: int,
    text: str,
    **kwargs,
) -> bool:
    try:
        await bot.send_message(
            user_id,
            text,
            **kwargs,
        )
        return True
    except TelegramForbiddenError:
        logger.warning("User blocked bot: %s", user_id)
        return False
    except TelegramBadRequest as exc:
        logger.warning("Telegram error for %s: %s", user_id, exc)
        return False
    except Exception:
        logger.exception("Unexpected send error")
        return False


# ============================================================
# KEYBOARDS
# ============================================================

def main_menu():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="🌐 Veb sayt bo'yicha ma'lumotlar"),
                KeyboardButton(text="🎨 Logo buyurtma qilish"),
            ],
            [
                KeyboardButton(text="ℹ️ Ma'lumot"),
                KeyboardButton(text="💡 Taklif va Shikoyatlar"),
            ],
            [
                KeyboardButton(text="🌐 Ijtimoiy tarmoqlar"),
                KeyboardButton(text="💬 Admin bilan muloqot"),
            ],
            [
                KeyboardButton(text="🧮 Loyiha kalkulyatori"),
                KeyboardButton(text="📂 Bizning portfolio"),
            ],
            [
                KeyboardButton(text="🎁 Aksiyalar va Referal"),
            ],
        ],
        resize_keyboard=True,
    )


def back_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="⬅️ Orqaga")],
        ],
        resize_keyboard=True,
    )


def info_menu():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="🤖 Bot haqida"),
                KeyboardButton(text="🏢 MHDV haqida"),
            ],
            [
                KeyboardButton(text="⬅️ Orqaga"),
            ],
        ],
        resize_keyboard=True,
    )


def feedback_menu():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="💡 Takliflar"),
                KeyboardButton(text="⚠️ Shikoyatlar"),
            ],
            [
                KeyboardButton(text="⬅️ Orqaga"),
            ],
        ],
        resize_keyboard=True,
    )


def skip_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Yo'q (O'tkazib yuborish)")],
            [KeyboardButton(text="⬅️ Orqaga")],
        ],
        resize_keyboard=True,
    )


def calculator_type_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="🌐 Veb sayt"),
                KeyboardButton(text="🎨 Logo"),
            ],
            [
                KeyboardButton(text="📱 Telegram bot"),
                KeyboardButton(text="🖥 Web tizim"),
            ],
            [
                KeyboardButton(text="⬅️ Orqaga"),
            ],
        ],
        resize_keyboard=True,
    )


def calculator_size_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="🟢 Oddiy"),
                KeyboardButton(text="🟡 O'rta"),
            ],
            [
                KeyboardButton(text="🔴 Murakkab"),
            ],
            [
                KeyboardButton(text="⬅️ Orqaga"),
            ],
        ],
        resize_keyboard=True,
    )


def admin_menu():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="👥 Foydalanuvchilar"),
                KeyboardButton(text="📊 Statistika"),
            ],
            [
                KeyboardButton(text="💬 Chatlar"),
                KeyboardButton(text="💡 Taklif va Shikoyatlar Boshqaruvi"),
            ],
            [
                KeyboardButton(text="📦 Zakazlar"),
                KeyboardButton(text="💳 To'lovlar"),
            ],
            [
                KeyboardButton(text="⚙️ Ma'lumotlarni almashtirish"),
                KeyboardButton(text="📢 Ommaviy xabar"),
            ],
            [
                KeyboardButton(text="🏷️ Promokod yaratish"),
                KeyboardButton(text="📥 Baza eksport"),
            ],
            [
                KeyboardButton(text="⬅️ Orqaga"),
            ],
        ],
        resize_keyboard=True,
    )


def order_admin_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="🌐 Veb sayt"),
                KeyboardButton(text="🎨 Logo"),
            ],
            [
                KeyboardButton(text="⬅️ Orqaga"),
            ],
        ],
        resize_keyboard=True,
    )


def content_admin_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Veb sayt ma'lumotlari")],
            [KeyboardButton(text="Logo ma'lumotlari")],
            [KeyboardButton(text="Bot haqida")],
            [KeyboardButton(text="MHDV haqida")],
            [KeyboardButton(text="Karta ma'lumotlari")],
            [KeyboardButton(text="Ijtimoiy tarmoqlar")],
            [KeyboardButton(text="⬅️ Orqaga")],
        ],
        resize_keyboard=True,
    )


def admin_order_action_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="✅ Qabul qilish")],
            [KeyboardButton(text="❌ Rad etish")],
            [KeyboardButton(text="⬅️ Orqaga")],
        ],
        resize_keyboard=True,
    )


# ============================================================
# FSM STATES
# ============================================================

class WebsiteOrderStates(StatesGroup):
    name = State()
    purpose = State()
    budget = State()
    requirements = State()
    additional = State()


class LogoOrderStates(StatesGroup):
    brand_name = State()
    niche = State()
    budget = State()
    additional = State()


class FeedbackStates(StatesGroup):
    waiting_text = State()


class ContactAdminStates(StatesGroup):
    waiting_text = State()


class AdminLoginStates(StatesGroup):
    password = State()


class AdminUserStates(StatesGroup):
    page = State()
    selected_user = State()
    message_text = State()
    discount = State()


class AdminOrderStates(StatesGroup):
    category = State()
    page = State()
    selected_order = State()
    rejection_reason = State()


class AdminPaymentStates(StatesGroup):
    selected_payment = State()


class AdminFeedbackStates(StatesGroup):
    selected_feedback = State()
    response = State()


class DynamicContentStates(StatesGroup):
    selected_key = State()
    new_value = State()


class BroadcastStates(StatesGroup):
    waiting_message = State()


class PromoStates(StatesGroup):
    code = State()
    discount = State()
    max_uses = State()
    expiry = State()


class CalculatorStates(StatesGroup):
    project_type = State()
    project_size = State()


# ============================================================
# ROUTER
# ============================================================

router = Router()


# ============================================================
# GLOBAL BLOCK CHECK
# ============================================================

@router.message()
async def blocked_guard(message: Message):
    if not message.from_user:
        return

    if await is_blocked_user(message.from_user.id):
        await message.answer(
            "🚫 Siz botdan foydalanish huquqidan mahrum qilingansiz."
        )
        return


# ============================================================
# START
# ============================================================

@router.message(Command("start"))
async def start_handler(
    message: Message,
    state: FSMContext,
):
    await state.clear()

    user = await ensure_user(message)

    if user is None:
        return

    first_name = (
        message.from_user.first_name
        if message.from_user
        else user.full_name
    )

    await message.answer(
        f"Salom {first_name}, MHDV botiga xush kelibsiz!",
        reply_markup=main_menu(),
    )


# ============================================================
# BACK BUTTON
# ============================================================

@router.message(F.text == "⬅️ Orqaga")
async def global_back(
    message: Message,
    state: FSMContext,
):
    current_state = await state.get_state()

    if current_state is None:
        await message.answer(
            "Asosiy menyu:",
            reply_markup=main_menu(),
        )
        return

    await state.clear()

    if message.from_user and message.from_user.id == ADMIN_ID:
        await message.answer(
            "Admin menyusi:",
            reply_markup=admin_menu(),
        )
    else:
        await message.answer(
            "Asosiy menyu:",
            reply_markup=main_menu(),
        )


# ============================================================
# WEBSITE ORDER
# ============================================================

@router.message(F.text == "🌐 Veb sayt bo'yicha ma'lumotlar")
async def website_start(
    message: Message,
    state: FSMContext,
):
    await state.clear()
    await state.set_state(WebsiteOrderStates.name)

    await message.answer(
        await get_dynamic_content("ws_info"),
        reply_markup=back_keyboard(),
    )

    await message.answer(
        "1/5. Ismingiz yoki tashkilot nomini kiriting:",
        reply_markup=back_keyboard(),
    )


@router.message(WebsiteOrderStates.name)
async def website_name(
    message: Message,
    state: FSMContext,
):
    await state.update_data(name=message.text)
    await state.set_state(WebsiteOrderStates.purpose)

    await message.answer(
        "2/5. Sayt qaysi maqsad yoki soha uchun kerak?",
        reply_markup=back_keyboard(),
    )


@router.message(WebsiteOrderStates.purpose)
async def website_purpose(
    message: Message,
    state: FSMContext,
):
    await state.update_data(purpose=message.text)
    await state.set_state(WebsiteOrderStates.budget)

    await message.answer(
        "3/5. Taxminiy budjetingiz qancha?",
        reply_markup=back_keyboard(),
    )


@router.message(WebsiteOrderStates.budget)
async def website_budget(
    message: Message,
    state: FSMContext,
):
    await state.update_data(budget=message.text)
    await state.set_state(WebsiteOrderStates.requirements)

    await message.answer(
        "4/5. Asosiy talablaringizni yozing:",
        reply_markup=back_keyboard(),
    )


@router.message(WebsiteOrderStates.requirements)
async def website_requirements(
    message: Message,
    state: FSMContext,
):
    await state.update_data(requirements=message.text)
    await state.set_state(WebsiteOrderStates.additional)

    await message.answer(
        "5/5. Qo'shimcha ma'lumot bo'lsa yozing:",
        reply_markup=skip_keyboard(),
    )


@router.message(WebsiteOrderStates.additional)
async def website_finish(
    message: Message,
    state: FSMContext,
    bot: Bot,
):
    additional = message.text

    if additional == "Yo'q (O'tkazib yuborish)":
        additional = "Qo'shimcha ma'lumot berilmagan."

    data = await state.get_data()

    details = (
        f"Ism/Tashkilot: {data.get('name')}\n"
        f"Soha/Maqsad: {data.get('purpose')}\n"
        f"Budjet: {data.get('budget')}\n"
        f"Asosiy talablar: {data.get('requirements')}\n"
        f"Qo'shimcha: {additional}"
    )

    async with SessionLocal() as session:
        order = Order(
            user_id=message.from_user.id,
            order_type="website",
            details=details,
            status="pending",
        )

        session.add(order)
        await session.commit()
        await session.refresh(order)

        order_id = order.id

    await state.clear()

    await message.answer(
        f"Arizangiz qabul qilindi!\n\n"
        f"Buyurtma ID: #{order_id}\n"
        f"Tez orada administrator siz bilan bog'lanadi.",
        reply_markup=main_menu(),
    )

    username = (
        f"@{message.from_user.username}"
        if message.from_user.username
        else "username yo'q"
    )

    admin_text = (
        "🆕 YANGI VEB SAYT BUYURTMASI\n\n"
        f"ID: #{order_id}\n"
        f"Foydalanuvchi: {message.from_user.full_name}\n"
        f"Username: {username}\n"
        f"Telegram ID: {message.from_user.id}\n\n"
        f"{details}"
    )

    await safe_send(
        bot,
        ADMIN_ID,
        admin_text,
    )


# ============================================================
# LOGO ORDER
# ============================================================

@router.message(F.text == "🎨 Logo buyurtma qilish")
async def logo_start(
    message: Message,
    state: FSMContext,
):
    await state.clear()
    await state.set_state(LogoOrderStates.brand_name)

    await message.answer(
        await get_dynamic_content("logo_info"),
        reply_markup=back_keyboard(),
    )

    await message.answer(
        "1/4. Brend nomini kiriting:",
        reply_markup=back_keyboard(),
    )


@router.message(LogoOrderStates.brand_name)
async def logo_brand(
    message: Message,
    state: FSMContext,
):
    await state.update_data(brand_name=message.text)
    await state.set_state(LogoOrderStates.niche)

    await message.answer(
        "2/4. Brend yo'nalishi yoki nishasini yozing:",
        reply_markup=back_keyboard(),
    )


@router.message(LogoOrderStates.niche)
async def logo_niche(
    message: Message,
    state: FSMContext,
):
    await state.update_data(niche=message.text)
    await state.set_state(LogoOrderStates.budget)

    await message.answer(
        "3/4. Taxminiy budjetingiz:",
        reply_markup=back_keyboard(),
    )


@router.message(LogoOrderStates.budget)
async def logo_budget(
    message: Message,
    state: FSMContext,
):
    await state.update_data(budget=message.text)
    await state.set_state(LogoOrderStates.additional)

    await message.answer(
        "4/4. Qo'shimcha ma'lumot:",
        reply_markup=skip_keyboard(),
    )


@router.message(LogoOrderStates.additional)
async def logo_finish(
    message: Message,
    state: FSMContext,
    bot: Bot,
):
    additional = message.text

    if additional == "Yo'q (O'tkazib yuborish)":
        additional = "Qo'shimcha ma'lumot berilmagan."

    data = await state.get_data()

    details = (
        f"Brend nomi: {data.get('brand_name')}\n"
        f"Yo'nalish: {data.get('niche')}\n"
        f"Budjet: {data.get('budget')}\n"
        f"Qo'shimcha: {additional}"
    )

    async with SessionLocal() as session:
        order = Order(
            user_id=message.from_user.id,
            order_type="logo",
            details=details,
            status="pending",
        )

        session.add(order)
        await session.commit()
        await session.refresh(order)

        order_id = order.id

    await state.clear()

    await message.answer(
        f"Arizangiz qabul qilindi!\n\n"
        f"Buyurtma ID: #{order_id}\n"
        f"Administrator tez orada siz bilan bog'lanadi.",
        reply_markup=main_menu(),
    )

    username = (
        f"@{message.from_user.username}"
        if message.from_user.username
        else "username yo'q"
    )

    admin_text = (
        "🆕 YANGI LOGO BUYURTMASI\n\n"
        f"ID: #{order_id}\n"
        f"Foydalanuvchi: {message.from_user.full_name}\n"
        f"Username: {username}\n"
        f"Telegram ID: {message.from_user.id}\n\n"
        f"{details}"
    )

    await safe_send(
        bot,
        ADMIN_ID,
        admin_text,
    )


# ============================================================
# INFORMATION
# ============================================================

@router.message(F.text == "ℹ️ Ma'lumot")
async def information_menu(message: Message):
    await message.answer(
        "Ma'lumot bo'limi:",
        reply_markup=info_menu(),
    )


@router.message(F.text == "🤖 Bot haqida")
async def about_bot(message: Message):
    await message.answer(
        await get_dynamic_content("about_bot"),
        reply_markup=info_menu(),
    )


@router.message(F.text == "🏢 MHDV haqida")
async def about_mhdv(message: Message):
    await message.answer(
        await get_dynamic_content("about_mhdv"),
        reply_markup=info_menu(),
    )


# ============================================================
# SOCIALS
# ============================================================

@router.message(F.text == "🌐 Ijtimoiy tarmoqlar")
async def socials(message: Message):
    await message.answer(
        await get_dynamic_content("socials"),
        reply_markup=main_menu(),
    )


# ============================================================
# FEEDBACK
# ============================================================

@router.message(F.text == "💡 Taklif va Shikoyatlar")
async def feedback_start(message: Message):
    await message.answer(
        "Kerakli bo'limni tanlang:",
        reply_markup=feedback_menu(),
    )


@router.message(F.text.in_(["💡 Takliflar", "⚠️ Shikoyatlar"]))
async def feedback_type(
    message: Message,
    state: FSMContext,
):
    feedback_type_value = (
        "proposal"
        if message.text == "💡 Takliflar"
        else "complaint"
    )

    await state.update_data(feedback_type=feedback_type_value)
    await state.set_state(FeedbackStates.waiting_text)

    title = (
        "taklifingizni"
        if feedback_type_value == "proposal"
        else "shikoyatingizni"
    )

    await message.answer(
        f"{title.capitalize()} batafsil yozing:",
        reply_markup=back_keyboard(),
    )


@router.message(FeedbackStates.waiting_text)
async def feedback_save(
    message: Message,
    state: FSMContext,
    bot: Bot,
):
    data = await state.get_data()

    async with SessionLocal() as session:
        feedback = Feedback(
            user_id=message.from_user.id,
            type=data["feedback_type"],
            text=message.text,
        )

        session.add(feedback)
        await session.commit()
        await session.refresh(feedback)

        feedback_id = feedback.id

    await state.clear()

    await message.answer(
        "Murojaatingiz qabul qilindi. Rahmat!",
        reply_markup=main_menu(),
    )

    feedback_label = (
        "Taklif"
        if data["feedback_type"] == "proposal"
        else "Shikoyat"
    )

    admin_text = (
        f"📩 YANGI {feedback_label.upper()}\n\n"
        f"ID: #{feedback_id}\n"
        f"Foydalanuvchi: {message.from_user.full_name}\n"
        f"Telegram ID: {message.from_user.id}\n"
        f"Username: @{message.from_user.username or 'yo‘q'}\n\n"
        f"{message.text}"
    )

    await safe_send(
        bot,
        ADMIN_ID,
        admin_text,
    )


# ============================================================
# CONTACT ADMIN
# ============================================================

@router.message(F.text == "💬 Admin bilan muloqot")
async def contact_admin_start(
    message: Message,
    state: FSMContext,
):
    await state.clear()
    await state.set_state(ContactAdminStates.waiting_text)

    await message.answer(
        "Administratorga yubormoqchi bo'lgan xabaringizni yozing:",
        reply_markup=back_keyboard(),
    )


@router.message(ContactAdminStates.waiting_text)
async def contact_admin_finish(
    message: Message,
    state: FSMContext,
    bot: Bot,
):
    await state.clear()

    username = (
        f"@{message.from_user.username}"
        if message.from_user.username
        else "username yo'q"
    )

    admin_text = (
        "💬 ADMIN UCHUN YANGI XABAR\n\n"
        f"Foydalanuvchi: {message.from_user.full_name}\n"
        f"Username: {username}\n"
        f"Telegram ID: {message.from_user.id}\n\n"
        f"{message.text}"
    )

    success = await safe_send(
        bot,
        ADMIN_ID,
        admin_text,
    )

    if success:
        await message.answer(
            "Xabaringiz administratorga yuborildi.",
            reply_markup=main_menu(),
        )
    else:
        await message.answer(
            "Xabarni yuborishda xatolik yuz berdi.",
            reply_markup=main_menu(),
        )


# ============================================================
# CALCULATOR
# ============================================================

@router.message(F.text == "🧮 Loyiha kalkulyatori")
async def calculator_start(
    message: Message,
    state: FSMContext,
):
    await state.clear()
    await state.set_state(CalculatorStates.project_type)

    await message.answer(
        "Loyiha turini tanlang:",
        reply_markup=calculator_type_keyboard(),
    )


@router.message(CalculatorStates.project_type)
async def calculator_type(
    message: Message,
    state: FSMContext,
):
    valid = [
        "🌐 Veb sayt",
        "🎨 Logo",
        "📱 Telegram bot",
        "🖥 Web tizim",
    ]

    if message.text not in valid:
        await message.answer(
            "Iltimos, menyudagi variantlardan birini tanlang.",
            reply_markup=calculator_type_keyboard(),
        )
        return

    await state.update_data(project_type=message.text)
    await state.set_state(CalculatorStates.project_size)

    await message.answer(
        "Loyiha murakkabligini tanlang:",
        reply_markup=calculator_size_keyboard(),
    )


@router.message(CalculatorStates.project_size)
async def calculator_size(
    message: Message,
    state: FSMContext,
):
    prices = {
        "🌐 Veb sayt": {
            "🟢 Oddiy": (800_000, 2_000_000),
            "🟡 O'rta": (2_000_000, 5_000_000),
            "🔴 Murakkab": (5_000_000, 12_000_000),
        },
        "🎨 Logo": {
            "🟢 Oddiy": (300_000, 700_000),
            "🟡 O'rta": (700_000, 1_500_000),
            "🔴 Murakkab": (1_500_000, 3_000_000),
        },
        "📱 Telegram bot": {
            "🟢 Oddiy": (800_000, 2_000_000),
            "🟡 O'rta": (2_000_000, 5_000_000),
            "🔴 Murakkab": (5_000_000, 10_000_000),
        },
        "🖥 Web tizim": {
            "🟢 Oddiy": (2_000_000, 5_000_000),
            "🟡 O'rta": (5_000_000, 12_000_000),
            "🔴 Murakkab": (12_000_000, 30_000_000),
        },
    }

    if message.text not in prices["🌐 Veb sayt"]:
        pass

    data = await state.get_data()
    project_type = data.get("project_type")

    if message.text not in prices.get(project_type, {}):
        await message.answer(
            "Iltimos, menyudagi variantlardan birini tanlang.",
            reply_markup=calculator_size_keyboard(),
        )
        return

    low, high = prices[project_type][message.text]

    await state.clear()

    await message.answer(
        "🧮 TAXMINI HISOB\n\n"
        f"Loyiha: {project_type}\n"
        f"Murakkablik: {message.text}\n\n"
        f"Taxminiy narx: {low:,} — {high:,} so'm\n\n"
        "Aniq narx loyiha talablari asosida belgilanadi.",
        reply_markup=main_menu(),
    )


# ============================================================
# PORTFOLIO
# ============================================================

@router.message(F.text == "📂 Bizning portfolio")
async def portfolio(message: Message):
    await message.answer(
        "📂 MHDV PORTFOLIO\n\n"
        "🌐 Web saytlar\n"
        "• Landing page\n"
        "• Portfolio saytlar\n"
        "• Korporativ saytlar\n"
        "• Internet do'konlar\n\n"
        "🎨 Logo dizaynlar\n"
        "• Minimal logo\n"
        "• Premium logo\n"
        "• Brand identity\n\n"
        "🖥 Web tizimlar\n"
        "• CRM\n"
        "• Admin panel\n"
        "• Buyurtma tizimlari\n\n"
        "Portfolio havolalari administrator tomonidan "
        "dinamik kontent orqali yangilanadi.",
        reply_markup=main_menu(),
    )


# ============================================================
# PROMOTIONS + REFERRAL
# ============================================================

@router.message(F.text == "🎁 Aksiyalar va Referal")
async def promotions(
    message: Message,
    bot: Bot,
):
    try:
        me = await bot.get_me()

        async with SessionLocal() as session:
            result = await session.execute(
                select(PromoCode).where(
                    PromoCode.expires_at >= datetime.utcnow(),
                    PromoCode.used_count < PromoCode.max_uses,
                )
            )

            promos = result.scalars().all()

        promo_text = "🎁 FAOL AKSIYALAR\n\n"

        if promos:
            for promo in promos:
                promo_text += (
                    f"🏷️ {promo.code}\n"
                    f"Chegirma: {promo.discount_percent}%\n"
                    f"Qolgan foydalanish: "
                    f"{promo.max_uses - promo.used_count}\n\n"
                )
        else:
            promo_text += "Hozircha faol promokod mavjud emas.\n\n"

        referral = (
            f"https://t.me/{me.username}"
            f"?start={message.from_user.id}"
        )

        promo_text += (
            "🔗 SIZNING REFERAL HAVOLANGIZ\n\n"
            f"{referral}\n\n"
            "Ushbu havolani do'stlaringizga yuborishingiz mumkin."
        )

        await message.answer(
            promo_text,
            reply_markup=main_menu(),
        )

    except Exception:
        logger.exception("Promotion error")

        await message.answer(
            "Aksiyalarni yuklashda xatolik yuz berdi.",
            reply_markup=main_menu(),
        )


# ============================================================
# ADMIN AUTH
# ============================================================

@router.message(Command("admin"))
async def admin_command(
    message: Message,
    state: FSMContext,
):
    if not message.from_user:
        return

    if message.from_user.id != ADMIN_ID:
        await message.answer(
            "🚫 Sizda admin panelga kirish huquqi yo'q."
        )
        return

    await state.clear()
    await state.set_state(AdminLoginStates.password)

    await message.answer(
        "🔐 Admin parolini kiriting:",
        reply_markup=back_keyboard(),
    )


@router.message(AdminLoginStates.password)
async def admin_password(
    message: Message,
    state: FSMContext,
):
    if message.from_user.id != ADMIN_ID:
        await state.clear()
        return

    entered = (message.text or "").strip().lower()

    if entered != ADMIN_PASSWORD:
        await message.answer(
            "❌ Parol noto'g'ri. Qayta kiriting:",
            reply_markup=back_keyboard(),
        )
        return

    await state.clear()

    await message.answer(
        "✅ Admin panelga xush kelibsiz.",
        reply_markup=admin_menu(),
    )


def admin_only(message: Message) -> bool:
    return bool(
        message.from_user
        and message.from_user.id == ADMIN_ID
    )


# ============================================================
# ADMIN STATISTICS
# ============================================================

@router.message(
    F.text == "📊 Statistika",
)
async def admin_statistics(message: Message):
    if not admin_only(message):
        return

    today_start = datetime.combine(
        datetime.utcnow().date(),
        datetime.min.time(),
    )

    async with SessionLocal() as session:
        total_users = (
            await session.scalar(
                select(func.count()).select_from(User)
            )
        ) or 0

        today_users = (
            await session.scalar(
                select(func.count())
                .select_from(User)
                .where(User.created_at >= today_start)
            )
        ) or 0

        blocked_users = (
            await session.scalar(
                select(func.count())
                .select_from(User)
                .where(User.is_blocked.is_(True))
            )
        ) or 0

        total_orders = (
            await session.scalar(
                select(func.count()).select_from(Order)
            )
        ) or 0

        approved = (
            await session.scalar(
                select(func.count())
                .select_from(Order)
                .where(Order.status == "approved")
            )
        ) or 0

        rejected = (
            await session.scalar(
                select(func.count())
                .select_from(Order)
                .where(Order.status == "rejected")
            )
        ) or 0

        pending = (
            await session.scalar(
                select(func.count())
                .select_from(Order)
                .where(Order.status == "pending")
            )
        ) or 0

        feedbacks = (
            await session.scalar(
                select(func.count()).select_from(Feedback)
            )
        ) or 0

    text = (
        "📊 REAL-TIME STATISTIKA\n\n"
        f"👥 Jami foydalanuvchilar: {total_users}\n"
        f"🆕 Bugungi yangi foydalanuvchilar: {today_users}\n"
        f"🚫 Bloklanganlar: {blocked_users}\n\n"
        f"📦 Jami zakazlar: {total_orders}\n"
        f"✅ Tasdiqlangan: {approved}\n"
        f"❌ Rad etilgan: {rejected}\n"
        f"⏳ Kutilayotgan: {pending}\n\n"
        f"💡 Feedbacklar: {feedbacks}"
    )

    await message.answer(
        text,
        reply_markup=admin_menu(),
    )


# ============================================================
# ADMIN USERS
# ============================================================

async def send_users_page(
    message: Message,
    page: int,
):
    per_page = 10

    async with SessionLocal() as session:
        total = (
            await session.scalar(
                select(func.count()).select_from(User)
            )
        ) or 0

        pages = max(1, math.ceil(total / per_page))

        page = max(0, min(page, pages - 1))

        result = await session.execute(
            select(User)
            .order_by(User.created_at.desc())
            .offset(page * per_page)
            .limit(per_page)
        )

        users = result.scalars().all()

    keyboard = []

    for user in users:
        status = "🚫" if user.is_blocked else "👤"

        keyboard.append(
            [
                KeyboardButton(
                    text=f"{status} {user.full_name[:35]} | {user.telegram_id}"
                )
            ]
        )

    navigation = []

    if page > 0:
        navigation.append(
            KeyboardButton(text="⬅️ Oldingi")
        )

    if page < pages - 1:
        navigation.append(
            KeyboardButton(text="Keyingi ➡️")
        )

    if navigation:
        keyboard.append(navigation)

    keyboard.append(
        [KeyboardButton(text="⬅️ Orqaga")]
    )

    markup = ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
    )

    await message.answer(
        f"👥 Foydalanuvchilar\n\n"
        f"Sahifa: {page + 1}/{pages}\n"
        f"Jami: {total}",
        reply_markup=markup,
    )

    return page


@router.message(F.text == "👥 Foydalanuvchilar")
async def admin_users(
    message: Message,
    state: FSMContext,
):
    if not admin_only(message):
        return

    await state.clear()
    await state.set_state(AdminUserStates.page)
    await state.update_data(page=0)

    await send_users_page(message, 0)


@router.message(
    AdminUserStates.page,
    F.text == "⬅️ Oldingi",
)
async def users_previous(
    message: Message,
    state: FSMContext,
):
    data = await state.get_data()
    page = max(0, int(data.get("page", 0)) - 1)

    await state.update_data(page=page)

    await send_users_page(message, page)


@router.message(
    AdminUserStates.page,
    F.text == "Keyingi ➡️",
)
async def users_next(
    message: Message,
    state: FSMContext,
):
    data = await state.get_data()

    current_page = int(data.get("page", 0))

    async with SessionLocal() as session:
        total = (
            await session.scalar(
                select(func.count()).select_from(User)
            )
        ) or 0

    pages = max(1, math.ceil(total / 10))

    page = min(current_page + 1, pages - 1)

    await state.update_data(page=page)

    await send_users_page(message, page)


@router.message(AdminUserStates.page)
async def select_admin_user(
    message: Message,
    state: FSMContext,
):
    text = message.text or ""

    try:
        telegram_id = int(text.rsplit("|", 1)[-1].strip())
    except ValueError:
        await message.answer(
            "Foydalanuvchini menyudan tanlang."
        )
        return

    async with SessionLocal() as session:
        result = await session.execute(
            select(User).where(
                User.telegram_id == telegram_id
            )
        )

        user = result.scalar_one_or_none()

        if not user:
            await message.answer("Foydalanuvchi topilmadi.")
            return

        order_count = (
            await session.scalar(
                select(func.count())
                .select_from(Order)
                .where(Order.user_id == telegram_id)
            )
        ) or 0

    await state.set_state(AdminUserStates.selected_user)
    await state.update_data(
        selected_user=telegram_id,
    )

    status = "Bloklangan" if user.is_blocked else "Faol"

    await message.answer(
        "👤 FOYDALANUVCHI\n\n"
        f"Ism: {user.full_name}\n"
        f"Telegram ID: {user.telegram_id}\n"
        f"Username: @{user.username or 'yo‘q'}\n"
        f"Zakazlar: {order_count}\n"
        f"Holat: {status}\n"
        f"Shaxsiy chegirma: {user.discount_percent}%",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="✉️ Bot orqali yozish")],
                [
                    KeyboardButton(
                        text="🚫 Blocklash / Unblocklash"
                    )
                ],
                [
                    KeyboardButton(
                        text="👤 Shaxsiy chatiga yozish"
                    )
                ],
                [
                    KeyboardButton(
                        text="📊 Foydalanuvchi xaridlari tarixi"
                    )
                ],
                [
                    KeyboardButton(
                        text="🏷️ Maxsus chegirma berish"
                    )
                ],
                [
                    KeyboardButton(text="⬅️ Orqaga")
                ],
            ],
            resize_keyboard=True,
        ),
    )


@router.message(
    AdminUserStates.selected_user,
    F.text == "🚫 Blocklash / Unblocklash",
)
async def toggle_user_block(
    message: Message,
    state: FSMContext,
):
    data = await state.get_data()
    telegram_id = int(data["selected_user"])

    async with SessionLocal() as session:
        result = await session.execute(
            select(User).where(
                User.telegram_id == telegram_id
            )
        )

        user = result.scalar_one_or_none()

        if not user:
            await message.answer("Foydalanuvchi topilmadi.")
            return

        user.is_blocked = not user.is_blocked

        await session.commit()

        status = (
            "bloklandi"
            if user.is_blocked
            else "blokdan chiqarildi"
        )

    await message.answer(
        f"Foydalanuvchi {status}.",
    )


@router.message(
    AdminUserStates.selected_user,
    F.text == "👤 Shaxsiy chatiga yozish",
)
async def private_chat_link(
    message: Message,
    state: FSMContext,
):
    data = await state.get_data()
    telegram_id = int(data["selected_user"])

    await message.answer(
        f"👤 Foydalanuvchi chatiga o'tish:\n\n"
        f'<a href="tg://user?id={telegram_id}">Shaxsiy chatni ochish</a>',
        parse_mode=ParseMode.HTML,
        reply_markup=message.reply_markup,
    )


@router.message(
    AdminUserStates.selected_user,
    F.text == "✉️ Bot orqali yozish",
)
async def admin_user_message_start(
    message: Message,
    state: FSMContext,
):
    await state.set_state(AdminUserStates.message_text)

    await message.answer(
        "Foydalanuvchiga yuboriladigan xabarni yozing:",
        reply_markup=back_keyboard(),
    )


@router.message(AdminUserStates.message_text)
async def admin_user_message_finish(
    message: Message,
    state: FSMContext,
    bot: Bot,
):
    data = await state.get_data()
    telegram_id = int(data["selected_user"])

    success = await safe_send(
        bot,
        telegram_id,
        f"📩 Administrator xabari:\n\n{message.text}",
    )

    await state.set_state(AdminUserStates.selected_user)

    await message.answer(
        "✅ Xabar yuborildi."
        if success
        else "❌ Xabarni yuborib bo'lmadi.",
    )


@router.message(
    AdminUserStates.selected_user,
    F.text == "🏷️ Maxsus chegirma berish",
)
async def discount_start(
    message: Message,
    state: FSMContext,
):
    await state.set_state(AdminUserStates.discount)

    await message.answer(
        "Chegirma foizini kiriting (0-100):",
        reply_markup=back_keyboard(),
    )


@router.message(AdminUserStates.discount)
async def discount_finish(
    message: Message,
    state: FSMContext,
):
    try:
        discount = int(message.text.strip())

        if not 0 <= discount <= 100:
            raise ValueError

    except (ValueError, AttributeError):
        await message.answer(
            "0 dan 100 gacha butun son kiriting."
        )
        return

    data = await state.get_data()
    telegram_id = int(data["selected_user"])

    async with SessionLocal() as session:
        result = await session.execute(
            select(User).where(
                User.telegram_id == telegram_id
            )
        )

        user = result.scalar_one_or_none()

        if user:
            user.discount_percent = discount
            await session.commit()

    await state.set_state(AdminUserStates.selected_user)

    await message.answer(
        f"✅ Shaxsiy chegirma {discount}% qilib belgilandi."
    )


@router.message(
    AdminUserStates.selected_user,
    F.text == "📊 Foydalanuvchi xaridlari tarixi",
)
async def user_order_history(
    message: Message,
    state: FSMContext,
):
    data = await state.get_data()
    telegram_id = int(data["selected_user"])

    async with SessionLocal() as session:
        result = await session.execute(
            select(Order)
            .where(Order.user_id == telegram_id)
            .order_by(Order.created_at.desc())
        )

        orders = result.scalars().all()

    if not orders:
        await message.answer(
            "Bu foydalanuvchida hali zakaz mavjud emas."
        )
        return

    text = "📊 FOYDALANUVCHI ZAKAZLARI\n\n"

    for order in orders:
        text += (
            f"#{order.id} | "
            f"{order.order_type} | "
            f"{order.status}\n"
            f"{order.created_at:%d.%m.%Y %H:%M}\n\n"
        )

    await message.answer(text)


# ============================================================
# ADMIN ORDER MANAGEMENT
# ============================================================

@router.message(F.text == "📦 Zakazlar")
async def admin_orders(
    message: Message,
    state: FSMContext,
):
    if not admin_only(message):
        return

    await state.clear()

    await message.answer(
        "Zakaz turini tanlang:",
        reply_markup=order_admin_keyboard(),
    )


async def send_pending_orders_page(
    message: Message,
    order_type: str,
    page: int,
):
    per_page = 10

    async with SessionLocal() as session:
        total = (
            await session.scalar(
                select(func.count())
                .select_from(Order)
                .where(
                    Order.order_type == order_type,
                    Order.status == "pending",
                )
            )
        ) or 0

        pages = max(1, math.ceil(total / per_page))
        page = max(0, min(page, pages - 1))

        result = await session.execute(
            select(Order)
            .where(
                Order.order_type == order_type,
                Order.status == "pending",
            )
            .order_by(Order.created_at.desc())
            .offset(page * per_page)
            .limit(per_page)
        )

        orders = result.scalars().all()

    keyboard = []

    for order in orders:
        keyboard.append(
            [
                KeyboardButton(
                    text=f"#{order.id} | {order.created_at:%d.%m.%Y}"
                )
            ]
        )

    navigation = []

    if page > 0:
        navigation.append(
            KeyboardButton(text="⬅️ Oldingi")
        )

    if page < pages - 1:
        navigation.append(
            KeyboardButton(text="Keyingi ➡️")
        )

    if navigation:
        keyboard.append(navigation)

    keyboard.append(
        [KeyboardButton(text="⬅️ Orqaga")]
    )

    await message.answer(
        f"📦 {order_type.upper()} ZAKAZLAR\n\n"
        f"Sahifa: {page + 1}/{pages}\n"
        f"Kutilayotgan: {total}",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=keyboard,
            resize_keyboard=True,
        ),
    )


@router.message(F.text.in_(["🌐 Veb sayt", "🎨 Logo"]))
async def admin_order_category(
    message: Message,
    state: FSMContext,
):
    if not admin_only(message):
        return

    order_type = (
        "website"
        if message.text == "🌐 Veb sayt"
        else "logo"
    )

    await state.set_state(AdminOrderStates.page)
    await state.update_data(
        category=order_type,
        page=0,
    )

    await send_pending_orders_page(
        message,
        order_type,
        0,
    )


@router.message(
    AdminOrderStates.page,
    F.text == "⬅️ Oldingi",
)
async def order_previous(
    message: Message,
    state: FSMContext,
):
    data = await state.get_data()

    page = max(
        0,
        int(data.get("page", 0)) - 1,
    )

    await state.update_data(page=page)

    await send_pending_orders_page(
        message,
        data["category"],
        page,
    )


@router.message(
    AdminOrderStates.page,
    F.text == "Keyingi ➡️",
)
async def order_next(
    message: Message,
    state: FSMContext,
):
    data = await state.get_data()

    order_type = data["category"]
    current_page = int(data.get("page", 0))

    async with SessionLocal() as session:
        total = (
            await session.scalar(
                select(func.count())
                .select_from(Order)
                .where(
                    Order.order_type == order_type,
                    Order.status == "pending",
                )
            )
        ) or 0

    pages = max(1, math.ceil(total / 10))

    page = min(current_page + 1, pages - 1)

    await state.update_data(page=page)

    await send_pending_orders_page(
        message,
        order_type,
        page,
    )


@router.message(AdminOrderStates.page)
async def select_order(
    message: Message,
    state: FSMContext,
):
    text = message.text or ""

    if not text.startswith("#"):
        return

    try:
        order_id = int(
            text.split("|")[0].replace("#", "").strip()
        )
    except ValueError:
        return

    async with SessionLocal() as session:
        result = await session.execute(
            select(Order).where(Order.id == order_id)
        )

        order = result.scalar_one_or_none()

        if not order:
            await message.answer("Zakaz topilmadi.")
            return

        user_result = await session.execute(
            select(User).where(
                User.telegram_id == order.user_id
            )
        )

        user = user_result.scalar_one_or_none()

    await state.set_state(AdminOrderStates.selected_order)
    await state.update_data(selected_order=order_id)

    text = (
        f"📦 BUYURTMA #{order.id}\n\n"
        f"Turi: {order.order_type}\n"
        f"Holati: {order.status}\n"
        f"Foydalanuvchi: "
        f"{user.full_name if user else 'Nomaʼlum'}\n"
        f"Telegram ID: {order.user_id}\n"
        f"Yaratilgan: {order.created_at:%d.%m.%Y %H:%M}\n\n"
        f"📋 BUYURTMA TAFSILOTLARI\n\n"
        f"{order.details}"
    )

    await message.answer(
        text,
        reply_markup=admin_order_action_keyboard(),
    )


# ============================================================
# APPROVE ORDER
# ============================================================

@router.message(
    AdminOrderStates.selected_order,
    F.text == "✅ Qabul qilish",
)
async def approve_order(
    message: Message,
    state: FSMContext,
    bot: Bot,
):
    data = await state.get_data()
    order_id = int(data["selected_order"])

    async with SessionLocal() as session:
        result = await session.execute(
            select(Order).where(Order.id == order_id)
        )

        order = result.scalar_one_or_none()

        if not order:
            await message.answer("Zakaz topilmadi.")
            return

        if order.status != "pending":
            await message.answer(
                "Bu zakaz allaqachon ko'rib chiqilgan."
            )
            return

        order.status = "approved"

        await session.commit()

        user_id = order.user_id

    card_info = await get_dynamic_content("card_info")

    payment_keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(
                    text="📞 Kontaktni yuborish"
                )
            ],
            [
                KeyboardButton(
                    text="⬅️ Orqaga"
                )
            ],
        ],
        resize_keyboard=True,
    )

    await safe_send(
        bot,
        user_id,
        "Xaridingiz tasdiqlandi! "
        "To'lov qilishingiz mumkin.\n\n"
        "Chek rasmini yoki PDF faylini yuboring.\n\n"
        f"{card_info}",
        reply_markup=payment_keyboard,
    )

    await state.clear()

    await message.answer(
        f"✅ #{order_id} zakaz tasdiqlandi.",
        reply_markup=admin_menu(),
    )


# ============================================================
# REJECT ORDER
# ============================================================

@router.message(
    AdminOrderStates.selected_order,
    F.text == "❌ Rad etish",
)
async def reject_order_start(
    message: Message,
    state: FSMContext,
):
    await state.set_state(
        AdminOrderStates.rejection_reason
    )

    await message.answer(
        "Rad etish sababini yozing:",
        reply_markup=back_keyboard(),
    )


@router.message(AdminOrderStates.rejection_reason)
async def reject_order_finish(
    message: Message,
    state: FSMContext,
    bot: Bot,
):
    data = await state.get_data()
    order_id = int(data["selected_order"])

    reason = message.text.strip()

    async with SessionLocal() as session:
        result = await session.execute(
            select(Order).where(Order.id == order_id)
        )

        order = result.scalar_one_or_none()

        if not order:
            await message.answer("Zakaz topilmadi.")
            return

        order.status = "rejected"
        order.rejection_reason = reason

        user_id = order.user_id

        await session.commit()

    await safe_send(
        bot,
        user_id,
        f"Xaridingiz tasdiqlanmadi. Sababi: {reason}",
    )

    await state.clear()

    await message.answer(
        f"❌ #{order_id} zakaz rad etildi.",
        reply_markup=admin_menu(),
    )


# ============================================================
# PAYMENT RECEIPTS
# ============================================================

@router.message(
    F.photo,
)
async def payment_photo_handler(
    message: Message,
    bot: Bot,
):
    if not message.from_user:
        return

    async with SessionLocal() as session:
        result = await session.execute(
            select(Order)
            .where(
                Order.user_id == message.from_user.id,
                Order.status == "approved",
            )
            .order_by(Order.created_at.desc())
        )

        order = result.scalars().first()

        if not order:
            await message.answer(
                "Sizda to'lov kutayotgan tasdiqlangan buyurtma topilmadi."
            )
            return

        payment = Payment(
            order_id=order.id,
            file_id=message.photo[-1].file_id,
            status="pending",
        )

        session.add(payment)

        await session.commit()
        await session.refresh(payment)

        payment_id = payment.id
        order_id = order.id

    await message.answer(
        "To'lov chekingiz qabul qilindi. "
        "Administrator tekshiradi."
    )

    await safe_send(
        bot,
        ADMIN_ID,
        "💳 YANGI TO'LOV CHEKI\n\n"
        f"To'lov ID: #{payment_id}\n"
        f"Zakaz ID: #{order_id}\n"
        f"Foydalanuvchi: {message.from_user.full_name}\n"
        f"Telegram ID: {message.from_user.id}",
    )

    try:
        await bot.send_photo(
            ADMIN_ID,
            message.photo[-1].file_id,
            caption=(
                f"💳 To'lov cheki #{payment_id}\n"
                f"Zakaz #{order_id}"
            ),
        )
    except Exception:
        logger.exception("Payment photo forwarding failed")


@router.message(F.document)
async def payment_document_handler(
    message: Message,
    bot: Bot,
):
    if not message.from_user:
        return

    document = message.document

    if document.mime_type not in (
        "application/pdf",
        "image/jpeg",
        "image/png",
    ):
        await message.answer(
            "Iltimos, to'lov chekini PDF yoki rasm formatida yuboring."
        )
        return

    async with SessionLocal() as session:
        result = await session.execute(
            select(Order)
            .where(
                Order.user_id == message.from_user.id,
                Order.status == "approved",
            )
            .order_by(Order.created_at.desc())
        )

        order = result.scalars().first()

        if not order:
            await message.answer(
                "Sizda to'lov kutayotgan tasdiqlangan buyurtma topilmadi."
            )
            return

        payment = Payment(
            order_id=order.id,
            file_id=document.file_id,
            status="pending",
        )

        session.add(payment)

        await session.commit()
        await session.refresh(payment)

        payment_id = payment.id
        order_id = order.id

    await message.answer(
        "To'lov hujjatingiz qabul qilindi. Administrator tekshiradi."
    )

    await safe_send(
        bot,
        ADMIN_ID,
        "💳 YANGI TO'LOV HUJJATI\n\n"
        f"To'lov ID: #{payment_id}\n"
        f"Zakaz ID: #{order_id}\n"
        f"Foydalanuvchi: {message.from_user.full_name}\n"
        f"Telegram ID: {message.from_user.id}",
    )

    try:
        await bot.send_document(
            ADMIN_ID,
            document.file_id,
            caption=(
                f"💳 To'lov hujjati #{payment_id}\n"
                f"Zakaz #{order_id}"
            ),
        )
    except Exception:
        logger.exception("Payment document forwarding failed")


@router.message(F.text == "📞 Kontaktni yuborish")
async def contact_button(
    message: Message,
):
    await message.answer(
        "Administrator siz bilan bog'lanadi.",
        reply_markup=main_menu(),
    )


# ============================================================
# ADMIN PAYMENTS
# ============================================================

@router.message(F.text == "💳 To'lovlar")
async def admin_payments(
    message: Message,
):
    if not admin_only(message):
        return

    async with SessionLocal() as session:
        result = await session.execute(
            select(Payment)
            .where(Payment.status == "pending")
            .order_by(Payment.created_at.desc())
            .limit(20)
        )

        payments = result.scalars().all()

    if not payments:
        await message.answer(
            "⏳ Kutilayotgan to'lovlar mavjud emas.",
            reply_markup=admin_menu(),
        )
        return

    keyboard = []

    for payment in payments:
        keyboard.append(
            [
                KeyboardButton(
                    text=f"💳 To'lov #{payment.id} | Zakaz #{payment.order_id}"
                )
            ]
        )

    keyboard.append(
        [KeyboardButton(text="⬅️ Orqaga")]
    )

    await message.answer(
        "💳 KUTILAYOTGAN TO'LOVLAR",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=keyboard,
            resize_keyboard=True,
        ),
    )


# ============================================================
# ADMIN FEEDBACK
# ============================================================

@router.message(
    F.text == "💡 Taklif va Shikoyatlar Boshqaruvi"
)
async def admin_feedbacks(
    message: Message,
):
    if not admin_only(message):
        return

    async with SessionLocal() as session:
        result = await session.execute(
            select(Feedback)
            .order_by(Feedback.created_at.desc())
            .limit(20)
        )

        feedbacks = result.scalars().all()

    if not feedbacks:
        await message.answer(
            "Feedbacklar mavjud emas.",
            reply_markup=admin_menu(),
        )
        return

    keyboard = []

    for feedback in feedbacks:
        label = (
            "💡"
            if feedback.type == "proposal"
            else "⚠️"
        )

        keyboard.append(
            [
                KeyboardButton(
                    text=f"{label} #{feedback.id}"
                )
            ]
        )

    keyboard.append(
        [KeyboardButton(text="⬅️ Orqaga")]
    )

    await message.answer(
        "💡 FEEDBACKLAR",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=keyboard,
            resize_keyboard=True,
        ),
    )


@router.message(
    F.text.regexp(r"^(💡|⚠️) #\d+$")
)
async def admin_feedback_select(
    message: Message,
    state: FSMContext,
):
    if not admin_only(message):
        return

    try:
        feedback_id = int(
            message.text.split("#")[1]
        )
    except (ValueError, IndexError):
        return

    async with SessionLocal() as session:
        result = await session.execute(
            select(Feedback)
            .where(Feedback.id == feedback_id)
        )

        feedback = result.scalar_one_or_none()

        if not feedback:
            await message.answer("Feedback topilmadi.")
            return

    await state.set_state(
        AdminFeedbackStates.selected_feedback
    )

    await state.update_data(
        selected_feedback=feedback_id,
        feedback_user=feedback.user_id,
    )

    label = (
        "Taklif"
        if feedback.type == "proposal"
        else "Shikoyat"
    )

    await message.answer(
        f"📩 {label}\n\n"
        f"ID: #{feedback.id}\n"
        f"User ID: {feedback.user_id}\n"
        f"Vaqt: {feedback.created_at:%d.%m.%Y %H:%M}\n\n"
        f"{feedback.text}\n\n"
        "Javob yozing:",
        reply_markup=back_keyboard(),
    )


@router.message(
    AdminFeedbackStates.selected_feedback
)
async def admin_feedback_response(
    message: Message,
    state: FSMContext,
    bot: Bot,
):
    data = await state.get_data()

    user_id = int(data["feedback_user"])

    success = await safe_send(
        bot,
        user_id,
        f"📩 Administrator javobi:\n\n{message.text}",
    )

    await state.clear()

    await message.answer(
        "✅ Javob yuborildi."
        if success
        else "❌ Javobni yuborib bo'lmadi.",
        reply_markup=admin_menu(),
    )


# ============================================================
# DYNAMIC CONTENT ADMIN
# ============================================================

CONTENT_KEYS = {
    "Veb sayt ma'lumotlari": "ws_info",
    "Logo ma'lumotlari": "logo_info",
    "Bot haqida": "about_bot",
    "MHDV haqida": "about_mhdv",
    "Karta ma'lumotlari": "card_info",
    "Ijtimoiy tarmoqlar": "socials",
}


@router.message(
    F.text == "⚙️ Ma'lumotlarni almashtirish"
)
async def content_management(
    message: Message,
):
    if not admin_only(message):
        return

    await message.answer(
        "Qaysi ma'lumotni o'zgartirmoqchisiz?",
        reply_markup=content_admin_keyboard(),
    )


@router.message(
    F.text.in_(list(CONTENT_KEYS.keys()))
)
async def content_select(
    message: Message,
    state: FSMContext,
):
    if not admin_only(message):
        return

    key = CONTENT_KEYS[message.text]

    current = await get_dynamic_content(key)

    await state.set_state(
        DynamicContentStates.new_value
    )

    await state.update_data(
        selected_key=key,
    )

    await message.answer(
        "Joriy ma'lumot:\n\n"
        f"{current}\n\n"
        "Yangi ma'lumotni yuboring:",
        reply_markup=back_keyboard(),
    )


@router.message(DynamicContentStates.new_value)
async def content_update(
    message: Message,
    state: FSMContext,
):
    data = await state.get_data()

    key = data["selected_key"]

    await set_dynamic_content(
        key,
        message.text,
    )

    await state.clear()

    await message.answer(
        "✅ Ma'lumot muvaffaqiyatli yangilandi.",
        reply_markup=admin_menu(),
    )


# ============================================================
# BROADCAST
# ============================================================

@router.message(F.text == "📢 Ommaviy xabar")
async def broadcast_start(
    message: Message,
    state: FSMContext,
):
    if not admin_only(message):
        return

    await state.set_state(
        BroadcastStates.waiting_message
    )

    await message.answer(
        "Barcha foydalanuvchilarga yuboriladigan xabarni "
        "yuboring.\n\n"
        "Matn, rasm yoki forward qilingan xabar yuborishingiz mumkin.",
        reply_markup=back_keyboard(),
    )


@router.message(BroadcastStates.waiting_message)
async def broadcast_finish(
    message: Message,
    state: FSMContext,
    bot: Bot,
):
    if not admin_only(message):
        return

    async with SessionLocal() as session:
        result = await session.execute(
            select(User.telegram_id)
        )

        user_ids = list(result.scalars().all())

    success = 0
    failed = 0

    for user_id in user_ids:
        try:
            if message.text:
                await bot.send_message(
                    user_id,
                    message.text,
                )

            elif message.photo:
                await bot.send_photo(
                    user_id,
                    message.photo[-1].file_id,
                    caption=message.caption,
                )

            elif message.document:
                await bot.send_document(
                    user_id,
                    message.document.file_id,
                    caption=message.caption,
                )

            elif message.video:
                await bot.send_video(
                    user_id,
                    message.video.file_id,
                    caption=message.caption,
                )

            else:
                await bot.copy_message(
                    chat_id=user_id,
                    from_chat_id=message.chat.id,
                    message_id=message.message_id,
                )

            success += 1

        except (
            TelegramForbiddenError,
            TelegramBadRequest,
        ):
            failed += 1

        except Exception:
            failed += 1
            logger.exception(
                "Broadcast error for %s",
                user_id,
            )

        await asyncio.sleep(0.04)

    await state.clear()

    await message.answer(
        "📢 BROADCAST YAKUNLANDI\n\n"
        f"✅ Muvaffaqiyatli: {success}\n"
        f"❌ Xatolik: {failed}\n"
        f"👥 Jami: {len(user_ids)}",
        reply_markup=admin_menu(),
    )


# ============================================================
# PROMO CODE
# ============================================================

@router.message(F.text == "🏷️ Promokod yaratish")
async def promo_start(
    message: Message,
    state: FSMContext,
):
    if not admin_only(message):
        return

    await state.clear()
    await state.set_state(PromoStates.code)

    await message.answer(
        "Promokod nomini kiriting:",
        reply_markup=back_keyboard(),
    )


@router.message(PromoStates.code)
async def promo_code(
    message: Message,
    state: FSMContext,
):
    code = message.text.strip().upper()

    if not code or len(code) > 100:
        await message.answer(
            "Promokod 1-100 belgidan iborat bo'lishi kerak."
        )
        return

    await state.update_data(code=code)
    await state.set_state(PromoStates.discount)

    await message.answer(
        "Chegirma foizini kiriting (1-100):",
        reply_markup=back_keyboard(),
    )


@router.message(PromoStates.discount)
async def promo_discount(
    message: Message,
    state: FSMContext,
):
    try:
        discount = int(message.text.strip())

        if not 1 <= discount <= 100:
            raise ValueError

    except (ValueError, AttributeError):
        await message.answer(
            "1-100 oralig'ida son kiriting."
        )
        return

    await state.update_data(
        discount=discount
    )

    await state.set_state(
        PromoStates.max_uses
    )

    await message.answer(
        "Maksimal foydalanish sonini kiriting:",
        reply_markup=back_keyboard(),
    )


@router.message(PromoStates.max_uses)
async def promo_max_uses(
    message: Message,
    state: FSMContext,
):
    try:
        max_uses = int(message.text.strip())

        if max_uses <= 0:
            raise ValueError

    except (ValueError, AttributeError):
        await message.answer(
            "Musbat butun son kiriting."
        )
        return

    await state.update_data(
        max_uses=max_uses
    )

    await state.set_state(
        PromoStates.expiry
    )

    await message.answer(
        "Amal qilish muddatini kiriting:\n"
        "Masalan: 31.12.2026",
        reply_markup=back_keyboard(),
    )


@router.message(PromoStates.expiry)
async def promo_expiry(
    message: Message,
    state: FSMContext,
):
    try:
        expires_at = datetime.strptime(
            message.text.strip(),
            "%d.%m.%Y",
        )

        expires_at = expires_at.replace(
            hour=23,
            minute=59,
            second=59,
        )

    except ValueError:
        await message.answer(
            "Sana formati noto'g'ri.\n"
            "Masalan: 31.12.2026"
        )
        return

    data = await state.get_data()

    async with SessionLocal() as session:
        existing = await session.execute(
            select(PromoCode).where(
                PromoCode.code == data["code"]
            )
        )

        promo = existing.scalar_one_or_none()

        if promo:
            await message.answer(
                "Bu promokod allaqachon mavjud."
            )
            return

        session.add(
            PromoCode(
                code=data["code"],
                discount_percent=data["discount"],
                max_uses=data["max_uses"],
                used_count=0,
                expires_at=expires_at,
            )
        )

        await session.commit()

    await state.clear()

    await message.answer(
        "✅ Promokod yaratildi.\n\n"
        f"Kod: {data['code']}\n"
        f"Chegirma: {data['discount']}%\n"
        f"Maksimal foydalanish: {data['max_uses']}\n"
        f"Amal qiladi: {expires_at:%d.%m.%Y}",
        reply_markup=admin_menu(),
    )


# ============================================================
# DATABASE EXPORT
# ============================================================

@router.message(F.text == "📥 Baza eksport")
async def database_export(
    message: Message,
    bot: Bot,
):
    if not admin_only(message):
        return

    try:
        async with SessionLocal() as session:
            users_result = await session.execute(
                select(User).order_by(User.created_at)
            )

            users = users_result.scalars().all()

            orders_result = await session.execute(
                select(Order).order_by(Order.created_at)
            )

            orders = orders_result.scalars().all()

            payments_result = await session.execute(
                select(Payment).order_by(Payment.created_at)
            )

            payments = payments_result.scalars().all()

        workbook = Workbook()

        ws_users = workbook.active
        ws_users.title = "Users"

        ws_users.append(
            [
                "Telegram ID",
                "Full Name",
                "Username",
                "Blocked",
                "Discount %",
                "Created At",
            ]
        )

        for user in users:
            ws_users.append(
                [
                    user.telegram_id,
                    user.full_name,
                    user.username,
                    user.is_blocked,
                    user.discount_percent,
                    user.created_at.strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),
                ]
            )

        ws_orders = workbook.create_sheet("Orders")

        ws_orders.append(
            [
                "ID",
                "User ID",
                "Type",
                "Details",
                "Status",
                "Rejection Reason",
                "Created At",
            ]
        )

        for order in orders:
            ws_orders.append(
                [
                    order.id,
                    order.user_id,
                    order.order_type,
                    order.details,
                    order.status,
                    order.rejection_reason,
                    order.created_at.strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),
                ]
            )

        ws_payments = workbook.create_sheet("Payments")

        ws_payments.append(
            [
                "ID",
                "Order ID",
                "File ID",
                "Status",
                "Created At",
            ]
        )

        for payment in payments:
            ws_payments.append(
                [
                    payment.id,
                    payment.order_id,
                    payment.file_id,
                    payment.status,
                    payment.created_at.strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),
                ]
            )

        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"

            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        max_length = max(
                            max_length,
                            len(str(cell.value or "")),
                        )
                    except Exception:
                        pass

                worksheet.column_dimensions[
                    column_letter
                ].width = min(max_length + 2, 60)

        buffer = io.BytesIO()

        workbook.save(buffer)

        buffer.seek(0)

        filename = (
            f"mhdv_database_"
            f"{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )

        document = BufferedInputFile(
            buffer.getvalue(),
            filename=filename,
        )

        await bot.send_document(
            message.chat.id,
            document,
            caption=(
                "📥 MHDV baza eksporti\n\n"
                f"Users: {len(users)}\n"
                f"Orders: {len(orders)}\n"
                f"Payments: {len(payments)}"
            ),
        )

    except Exception:
        logger.exception("Database export error")

        await message.answer(
            "Baza eksport qilishda xatolik yuz berdi.",
            reply_markup=admin_menu(),
        )


# ============================================================
# ADMIN CHAT LOG / RECENT ROUTED MESSAGES
# ============================================================

@router.message(F.text == "💬 Chatlar")
async def admin_chats(
    message: Message,
):
    if not admin_only(message):
        return

    await message.answer(
        "💬 CHAT TIZIMI\n\n"
        "Foydalanuvchilarning administratorga yuborgan "
        "xabarlarini olish uchun ular "
        "«💬 Admin bilan muloqot» bo'limidan foydalanadi.\n\n"
        "To'g'ridan-to'g'ri javob berish uchun "
        "«👥 Foydalanuvchilar» bo'limidan foydalanuvchini tanlang.",
        reply_markup=admin_menu(),
    )


# ============================================================
# ADMIN PAYMENT VERIFICATION BY TEXT BUTTON
# ============================================================

@router.message(
    F.text.regexp(r"^💳 To'lov #\d+ \| Zakaz #\d+$")
)
async def select_payment(
    message: Message,
    state: FSMContext,
):
    if not admin_only(message):
        return

    try:
        payment_id = int(
            message.text.split("#")[1].split("|")[0].strip()
        )
    except (ValueError, IndexError):
        await message.answer("To'lov topilmadi.")
        return

    async with SessionLocal() as session:
        result = await session.execute(
            select(Payment).where(
                Payment.id == payment_id
            )
        )

        payment = result.scalar_one_or_none()

        if not payment:
            await message.answer("To'lov topilmadi.")
            return

        order_result = await session.execute(
            select(Order).where(
                Order.id == payment.order_id
            )
        )

        order = order_result.scalar_one_or_none()

    await state.set_state(
        AdminPaymentStates.selected_payment
    )

    await state.update_data(
        selected_payment=payment_id
    )

    await message.answer(
        "💳 TO'LOV\n\n"
        f"To'lov ID: #{payment.id}\n"
        f"Zakaz ID: #{payment.order_id}\n"
        f"Status: {payment.status}\n"
        f"Yuborilgan: {payment.created_at:%d.%m.%Y %H:%M}\n\n"
        "To'lovni tekshirish uchun yuqoridagi "
        "chek faylini ko'rib chiqing.",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[
                [
                    KeyboardButton(
                        text="✅ To'lovni tasdiqlash"
                    )
                ],
                [
                    KeyboardButton(
                        text="❌ To'lovni rad etish"
                    )
                ],
                [
                    KeyboardButton(text="⬅️ Orqaga")
                ],
            ],
            resize_keyboard=True,
        ),
    )

    try:
        if payment.file_id:
            if payment.file_id.startswith("AgAC"):
                await message.bot.send_photo(
                    ADMIN_ID,
                    payment.file_id,
                    caption=f"To'lov #{payment.id}",
                )
            else:
                await message.bot.send_document(
                    ADMIN_ID,
                    payment.file_id,
                    caption=f"To'lov #{payment.id}",
                )
    except Exception:
        logger.exception("Could not display payment")


@router.message(
    AdminPaymentStates.selected_payment,
    F.text == "✅ To'lovni tasdiqlash",
)
async def verify_payment(
    message: Message,
    state: FSMContext,
    bot: Bot,
):
    data = await state.get_data()

    payment_id = int(data["selected_payment"])

    async with SessionLocal() as session:
        result = await session.execute(
            select(Payment).where(
                Payment.id == payment_id
            )
        )

        payment = result.scalar_one_or_none()

        if not payment:
            await message.answer("To'lov topilmadi.")
            return

        if payment.status == "verified":
            await message.answer(
                "Bu to'lov allaqachon tasdiqlangan."
            )
            return

        payment.status = "verified"

        order_result = await session.execute(
            select(Order).where(
                Order.id == payment.order_id
            )
        )

        order = order_result.scalar_one_or_none()

        if not order:
            await message.answer("Zakaz topilmadi.")
            return

        user_id = order.user_id

        await session.commit()

    await safe_send(
        bot,
        user_id,
        "To'lovingiz tasdiqlandi! "
        "Sizga admin aloqaga chiqadi.",
        reply_markup=main_menu(),
    )

    await state.clear()

    await message.answer(
        f"✅ To'lov #{payment_id} tasdiqlandi.",
        reply_markup=admin_menu(),
    )


@router.message(
    AdminPaymentStates.selected_payment,
    F.text == "❌ To'lovni rad etish",
)
async def reject_payment(
    message: Message,
    state: FSMContext,
):
    data = await state.get_data()

    payment_id = int(data["selected_payment"])

    async with SessionLocal() as session:
        result = await session.execute(
            select(Payment).where(
                Payment.id == payment_id
            )
        )

        payment = result.scalar_one_or_none()

        if payment:
            payment.status = "rejected"
            await session.commit()

    await state.clear()

    await message.answer(
        f"❌ To'lov #{payment_id} rad etildi.",
        reply_markup=admin_menu(),
    )


# ============================================================
# USER UNKNOWN COMMAND / TEXT
# ============================================================

@router.message()
async def fallback_handler(
    message: Message,
):
    if not message.from_user:
        return

    if message.from_user.id == ADMIN_ID:
        await message.answer(
            "Admin menyusidan foydalaning.",
            reply_markup=admin_menu(),
        )
    else:
        await message.answer(
            "Kerakli bo'limni menyudan tanlang.",
            reply_markup=main_menu(),
        )


# ============================================================
# ERROR HANDLER
# ============================================================

@router.errors()
async def global_error_handler(
    event,
):
    logger.exception(
        "Unhandled Telegram update error: %s",
        event.exception,
    )

    return True


# ============================================================
# BOT STARTUP
# ============================================================

async def main():
    await init_db()

    bot = Bot(
    token=BOT_TOKEN,
    default=DefaultBotProperties(
        parse_mode=ParseMode.HTML
    )
)

    dp = Dispatcher()

    dp.include_router(router)

    try:
        bot_info = await bot.get_me()

        logger.info(
            "Bot started: @%s",
            bot_info.username,
        )

        await dp.start_polling(
            bot,
            allowed_updates=dp.resolve_used_update_types(),
        )

    finally:
        await bot.session.close()
        await engine.dispose()


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Bot stopped by user.")
    except Exception:
        logger.exception("Fatal application error.")